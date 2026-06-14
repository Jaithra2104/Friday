"""
file_analyzer.py — Analyze files with AI for Friday.

Supports:
  - PDF (.pdf)           — pdfplumber
  - Word (.docx, .doc)   — python-docx
  - Text (.txt, .md, etc)
  - Excel (.xlsx, .xls)  — openpyxl
  - Images (.png, .jpg)  — Gemini Vision API
  - CSV (.csv)

Returns:
  - AI summary + key points
  - List of detected dates/times/deadlines for reminder prompting
"""

import os
import re
import base64
import json
import requests
from pathlib import Path
from urllib.parse import quote_plus


# ── Text extractors ───────────────────────────────────────────────────────

def _read_pdf(filepath):
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except ImportError:
        print("[FILE] pdfplumber not installed, trying PyPDF2...")
        try:
            import PyPDF2
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                return "\n".join(
                    page.extract_text() for page in reader.pages
                    if page.extract_text()
                )
        except Exception as e:
            print(f"[FILE] PDF read error: {e}")
            return None


def _read_docx(filepath):
    try:
        from docx import Document
        doc = Document(filepath)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        print(f"[FILE] DOCX read error: {e}")
        return None


def _read_text(filepath):
    for enc in ["utf-8", "cp1252", "latin-1"]:
        try:
            with open(filepath, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    return None


def _read_excel(filepath):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append(" | ".join(cells))
        return "\n".join(lines)
    except Exception as e:
        print(f"[FILE] Excel read error: {e}")
        return None


def _read_csv(filepath):
    import csv
    rows = []
    try:
        with open(filepath, newline="", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(" | ".join(row))
        return "\n".join(rows[:200])   # limit rows
    except Exception as e:
        print(f"[FILE] CSV read error: {e}")
        return None


def _image_to_base64(filepath):
    with open(filepath, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


# ── Main extractor ────────────────────────────────────────────────────────

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
TEXT_EXTS  = {".txt", ".md", ".py", ".js", ".html", ".json", ".xml",
              ".log", ".ini", ".cfg", ".yaml", ".yml"}

def extract_content(filepath):
    """
    Extract content from any supported file.
    Returns: (content, mode)
      mode = 'text' | 'image' | None
    """
    ext = Path(filepath).suffix.lower()

    if ext == ".pdf":
        return _read_pdf(filepath), "text"
    elif ext in (".docx", ".doc"):
        return _read_docx(filepath), "text"
    elif ext in (".xlsx", ".xls"):
        return _read_excel(filepath), "text"
    elif ext == ".csv":
        return _read_csv(filepath), "text"
    elif ext in TEXT_EXTS:
        return _read_text(filepath), "text"
    elif ext in IMAGE_EXTS:
        return _image_to_base64(filepath), "image"
    else:
        # Try plain text as a last resort
        content = _read_text(filepath)
        return content, "text" if content else None


# ── Timing / deadline detection ───────────────────────────────────────────

_MONTHS = (
    "january|february|march|april|may|june|july|august|"
    "september|october|november|december|"
    "jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec"
)

DATE_REGEXES = [
    # DD/MM/YYYY or MM/DD/YYYY
    re.compile(r'\b\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}\b'),
    # YYYY-MM-DD
    re.compile(r'\b\d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2}\b'),
    # Month DD, YYYY  or  DD Month YYYY
    re.compile(rf'\b(?:{_MONTHS})\.?\s+\d{{1,2}}(?:st|nd|rd|th)?,?\s+\d{{4}}\b', re.I),
    re.compile(rf'\b\d{{1,2}}(?:st|nd|rd|th)?\s+(?:{_MONTHS}),?\s+\d{{4}}\b', re.I),
    # "next Monday", "this Friday"
    re.compile(r'\b(?:next|this)\s+(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday)\b', re.I),
    # tomorrow, today
    re.compile(r'\b(?:tomorrow|today|tonight)\b', re.I),
]

TIME_REGEXES = [
    re.compile(r'\b\d{1,2}:\d{2}\s*(?:am|pm|AM|PM)?\b'),
    re.compile(r'\b\d{1,2}\s*(?:am|pm|AM|PM)\b'),
    re.compile(r'\b(?:noon|midnight)\b', re.I),
]

DEADLINE_KEYWORDS = [
    "deadline", "due by", "due on", "due date", "submit by",
    "meeting at", "meeting on", "appointment", "scheduled for",
    "interview", "exam on", "test on", "presentation on",
    "project due", "report due", "call at", "call on",
    "conference", "event on", "by end of", "must be", "no later than",
    "reminder", "don't forget", "by the", "before the"
]


def find_timings(text: str) -> list:
    """
    Scan text for date/time/deadline mentions.
    Returns list of dicts: {line, type, context}
    """
    findings = []
    seen_lines = set()
    lines = text.split("\n")

    for line in lines:
        stripped = line.strip()
        if not stripped or len(stripped) < 8:
            continue

        has_date    = any(rx.search(stripped) for rx in DATE_REGEXES)
        has_time    = any(rx.search(stripped) for rx in TIME_REGEXES)
        has_keyword = any(kw in stripped.lower() for kw in DEADLINE_KEYWORDS)

        if (has_date or has_time) and stripped not in seen_lines:
            seen_lines.add(stripped)
            findings.append({
                "line":       stripped[:200],
                "has_date":   has_date,
                "has_time":   has_time,
                "has_keyword": has_keyword,
                "priority":   3 if (has_keyword and has_date) else
                              2 if has_keyword else
                              1
            })

    # Sort by priority, limit to 5
    findings.sort(key=lambda x: -x["priority"])
    return findings[:5]


# ── AI analysis ───────────────────────────────────────────────────────────

ANALYSIS_SYSTEM_PROMPT = (
    "You are Friday, an AI assistant. Analyze the provided document content thoroughly. "
    "Respond with:\n"
    "1. SUMMARY: A 2-3 sentence overview.\n"
    "2. KEY POINTS: Bullet points of the most important information.\n"
    "3. DATES & DEADLINES: List any dates, times, or deadlines mentioned.\n"
    "4. ACTION ITEMS: Any tasks or follow-ups the user should take.\n"
    "Keep the spoken summary short (2-3 sentences). The full analysis will be shown on screen."
)

IMAGE_ANALYSIS_PROMPT = (
    "You are Friday, an AI assistant. Analyze this image thoroughly. "
    "Describe what you see, extract any text visible, note important details, "
    "and identify any dates, times, or important information. "
    "Respond with a structured analysis."
)


def analyze_with_gemini_vision(image_b64: str, mime_type: str,
                                api_key: str) -> str:
    """Send image to Gemini Vision for analysis."""
    url = ("https://generativelanguage.googleapis.com/v1beta/models/"
           "gemini-2.0-flash:generateContent")
    payload = {
        "contents": [{
            "parts": [
                {"text": IMAGE_ANALYSIS_PROMPT},
                {"inline_data": {"mime_type": mime_type, "data": image_b64}}
            ]
        }]
    }
    resp = requests.post(url, params={"key": api_key}, json=payload, timeout=30)
    resp.raise_for_status()
    parts = resp.json()["candidates"][0]["content"]["parts"]
    return " ".join(p.get("text", "") for p in parts).strip()


def get_short_summary(full_analysis: str) -> str:
    """Extract just the SUMMARY section for TTS (keep it short)."""
    lines = full_analysis.split("\n")
    summary_lines = []
    in_summary = False
    for line in lines:
        if "SUMMARY" in line.upper():
            in_summary = True
            continue
        if in_summary:
            if any(k in line.upper() for k in ["KEY POINT", "DATE", "ACTION", "##", "**"]):
                break
            if line.strip():
                summary_lines.append(line.strip())
    if summary_lines:
        return " ".join(summary_lines[:3])
    # Fallback: first 2 sentences
    sentences = re.split(r'(?<=[.!?])\s+', full_analysis)
    return " ".join(sentences[:2])
